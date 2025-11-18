#!/usr/bin/env python3
import sys
import csv
import os
import xml.etree.ElementTree as ET
from rdkit import Chem
from rdkit.Chem import AllChem, MACCSkeys, rdmolfiles
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

def is_csv_name(arg):
    """Return True if argument is a CSV filename."""
    return arg.lower().endswith(".csv")

def count_bits(bits_tuple):
    """Count how many bits are set to 1 in the MACCS fingerprint."""
    return sum(bits_tuple)

def parse_cdxml_to_mol(cdxml_file):
    """Parse CDXML file and create an RDKit molecule."""
    try:
        tree = ET.parse(cdxml_file)
        root = tree.getroot()
        
        # Find all nodes (atoms) and bonds
        nodes = {}
        bonds = []
        
        # Parse nodes (atoms)
        for node in root.iter('{http://www.cambridgesoft.com/xml/cdxml.dtd}n'):
            node_id = node.get('id')
            element = node.get('Element', '6')  # Default to carbon (atomic number 6)
            # If no Element attribute, it's likely carbon
            nodes[node_id] = element
        
        # If we find nodes without namespace, try without namespace
        if not nodes:
            for node in root.iter('n'):
                node_id = node.get('id')
                element = node.get('Element', '6')  # Default to carbon
                nodes[node_id] = element
        
        # Parse bonds
        for bond in root.iter('{http://www.cambridgesoft.com/xml/cdxml.dtd}b'):
            begin = bond.get('B')
            end = bond.get('E')
            order = bond.get('Order', '1')  # Default to single bond
            bonds.append((begin, end, order))
        
        # If we find bonds without namespace, try without namespace
        if not bonds:
            for bond in root.iter('b'):
                begin = bond.get('B')
                end = bond.get('E')
                order = bond.get('Order', '1')
                bonds.append((begin, end, order))
        
        if not nodes:
            return None
        
        # Create RDKit molecule
        mol = Chem.RWMol()
        atom_map = {}
        
        # Add atoms
        for node_id, element in nodes.items():
            try:
                atomic_num = int(element)
            except:
                atomic_num = 6  # Default to carbon
            atom = Chem.Atom(atomic_num)
            idx = mol.AddAtom(atom)
            atom_map[node_id] = idx
        
        # Add bonds
        for begin, end, order in bonds:
            if begin in atom_map and end in atom_map:
                try:
                    bond_order = int(float(order))
                    if bond_order == 1:
                        bond_type = Chem.BondType.SINGLE
                    elif bond_order == 2:
                        bond_type = Chem.BondType.DOUBLE
                    elif bond_order == 3:
                        bond_type = Chem.BondType.TRIPLE
                    else:
                        bond_type = Chem.BondType.SINGLE
                    mol.AddBond(atom_map[begin], atom_map[end], bond_type)
                except:
                    mol.AddBond(atom_map[begin], atom_map[end], Chem.BondType.SINGLE)
        
        # Convert to regular molecule and add hydrogens
        mol = mol.GetMol()
        mol_with_h = None
        try:
            Chem.SanitizeMol(mol)
            mol_with_h = Chem.AddHs(mol)
        except:
            # Try sanitizing without strict valence checking
            try:
                Chem.SanitizeMol(mol, sanitizeOps=Chem.SANITIZE_ALL^Chem.SANITIZE_PROPERTIES)
                mol_with_h = Chem.AddHs(mol)
            except Exception as e:
                print(f"Warning: Sanitization issues in {cdxml_file}: {e}")
                # Continue with molecule without added H
                mol_with_h = mol
        
        # Return molecule with H if successful, otherwise without
        return mol_with_h if mol_with_h is not None else mol
        
    except Exception as e:
        print(f"Error parsing {cdxml_file}: {e}")
        return None

def main():
    args = sys.argv[1:]
    if not args:
        print("Usage: python cdxml2csv.py *.cdxml [output.csv]")
        sys.exit(1)

    # Default CSV output name
    output_csv = "cdxml2csv.csv"

    # If last argument ends with .csv AND does not exist -> use it
    if is_csv_name(args[-1]) and not os.path.exists(args[-1]):
        output_csv = args[-1]
        cdxml_files = args[:-1]
    else:
        cdxml_files = args

    # Filter only CDXML files
    cdxml_files = [f for f in cdxml_files if f.lower().endswith(".cdxml")]

    if not cdxml_files:
        print("No .cdxml files provided.")
        sys.exit(1)

    # Open CSV for writing
    with open(output_csv, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile, delimiter=";")
        writer.writerow(["name", "smiles", "maccs"])

        max_bits = -1
        max_file = None

        # Process each CDXML file
        for filename in cdxml_files:
            mol = parse_cdxml_to_mol(filename)
            if mol is None:
                print(f"Could not parse {filename}")
                continue

            # Extract SMILES (remove hydrogens first for canonical SMILES)
            try:
                mol_no_h = Chem.RemoveHs(mol)
                smiles = Chem.MolToSmiles(mol_no_h)
            except Exception as e:
                # If removing H fails, try generating SMILES from original mol without explicit H in output
                try:
                    # For molecules with weird valence, use the allHsExplicit=False to suppress them in SMILES
                    mol_copy = Chem.Mol(mol)
                    for atom in mol_copy.GetAtoms():
                        if atom.GetAtomicNum() == 1:  # If it's hydrogen
                            continue
                    smiles = Chem.MolToSmiles(mol, allHsExplicit=False)
                    print(f"Note: Using non-canonical SMILES for {filename}")
                except:
                    print(f"Warning: Could not generate SMILES for {filename}: {e}")
                    smiles = "N/A"

            # MACCS fingerprint
            try:
                fp = MACCSkeys.GenMACCSKeys(mol)
                # Get list of bit indices that are set to 1 (similar to pybel's .bits)
                bits_list = [i for i, bit in enumerate(fp.ToBitString()) if bit == '1']
                bit_count = len(bits_list)
            except Exception as e:
                print(f"Could not calculate MACCS for {filename}: {e}")
                continue

            # Track molecule with highest bit count
            if bit_count > max_bits:
                max_bits = bit_count
                max_file = filename

            # Write CSV row - format bits_list as a string representation
            writer.writerow([
                filename,
                smiles,
                str(bits_list)
            ])

    print(f"Molecule with the most MACCS bits: {max_file} ({max_bits} bits set)")
    
    # Create formatted Excel file
    create_excel_file(output_csv)

def create_excel_file(csv_file):
    """Create a formatted Excel file from the CSV."""
    xlsx_file = csv_file.replace('.csv', '.xlsx')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "CDXML Data"
    
    # Read CSV and write to Excel
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Define border styles
    left_border = Border(left=Side(style='thin'))
    right_border = Border(right=Side(style='thin'))
    left_right_border = Border(left=Side(style='thin'), right=Side(style='thin'))
    bottom_right_border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    bottom_left_right_border = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    bottom_left_border = Border(bottom=Side(style='thin'), left=Side(style='thin'))
    
    # Format header row (bold and borders)
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=1).border = bottom_right_border
    ws.cell(row=1, column=2).font = Font(bold=True)
    ws.cell(row=1, column=2).border = bottom_left_right_border
    ws.cell(row=1, column=3).font = Font(bold=True)
    ws.cell(row=1, column=3).border = bottom_left_border
    
    # Add column borders to data cells
    max_row = ws.max_row
    for row_idx in range(2, max_row + 1):
        ws.cell(row=row_idx, column=1).border = right_border
        ws.cell(row=row_idx, column=2).border = left_right_border
        ws.cell(row=row_idx, column=3).border = left_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 80
    
    # Save the workbook
    wb.save(xlsx_file)
    print(f"Formatted Excel file created: {xlsx_file}")

if __name__ == "__main__":
    main()
